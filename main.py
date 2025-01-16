"""
DNS Resolver Benchmark Tool

This script performs performance testing of DNS-over-HTTPS (DoH) servers by:
- Reading a list of domains and DNS servers from configuration files
- Resolving each domain using each DNS server multiple times
- Measuring response times and error rates
- Generating detailed performance reports in Excel format

Requirements:
- dnslookup.exe in the same directory
- openpyxl package for Excel report generation
- colorama for console output formatting
"""

import os
import subprocess
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from itertools import cycle
from typing import List, Dict
import re

from colorama import Fore, Style
import pandas as pd

try:
    import importlib.util

    if importlib.util.find_spec("openpyxl") is None:
        print(
            "openpyxl module not found. Please install it using: pip install openpyxl"
        )
        raise ImportError(
            "The 'openpyxl' module is required for this script to run. Please install it using: pip install openpyxl"
        )
except ImportError as e:
    if "openpyxl" in str(e):
        print("Error importing openpyxl module.")
        raise ImportError("Error importing openpyxl module.")
    else:
        raise

from openpyxl.styles import Font, PatternFill, Border, Side

CONFIG = {
    "NUM_QUERIES": 30,  # Number of DNS queries to perform per domain-server pair
    "DOMAINS_FILE": "test_domains.txt",  # File containing list of domains to test
    "DNS_SERVERS_FILE": "dns_servers.txt",  # File containing list of DNS servers
    "DNSLOOKUP_EXE": "dnslookup.exe",  # Path to dnslookup executable
    "MAX_THREADS": 1,  # Maximum number of concurrent DNS resolution threads
}


class ColorFormatter:
    """
    Utility class for formatting console output with colors using colorama.
    Provides methods for success (green), error (red), and highlight (yellow) formatting.
    """
    @staticmethod
    def success(text: str) -> str:
        return f"{Fore.GREEN}{text}{Style.RESET_ALL}"

    @staticmethod
    def error(text: str) -> str:
        return f"{Fore.RED}{text}{Style.RESET_ALL}"

    @staticmethod
    def highlight(text: str) -> str:
        return f"{Fore.YELLOW}{text}{Style.RESET_ALL}"


class ProgressIndicator:
    """
    Displays an animated progress indicator in the console showing the number of completed tasks.
    Uses a spinning animation to indicate active processing.
    """
    def __init__(self, total_tasks: int):
        """
        Initialize the progress indicator.
        
        Args:
            total_tasks: Total number of tasks to be completed
        """
        self.total_tasks = total_tasks
        self.completed_tasks = 0
        self.running = False
        self.spinner = cycle(["⠋", "⠙", "⠹", "⠸", "⠼", "⠴", "⠦", "⠧", "⠇", "⠏"])
        self.thread = None

    def _spin(self):
        while self.running:
            progress = f"[{self.completed_tasks}/{self.total_tasks}]"
            sys.stdout.write(f"\rProcessing {progress} {next(self.spinner)} ")
            sys.stdout.flush()
            time.sleep(0.1)

    def start(self):
        self.running = True
        self.thread = threading.Thread(target=self._spin)
        self.thread.start()

    def stop(self):
        self.running = False
        if self.thread:
            self.thread.join()
        sys.stdout.write("\r" + " " * 50 + "\r")
        sys.stdout.flush()

    def update(self):
        self.completed_tasks += 1


class DNSResolver:
    """
    Handles DNS resolution using the dnslookup executable.
    Performs multiple queries and collects statistics about response times and errors.
    """
    def __init__(self, dnslookup_path: str):
        self.dnslookup_path = dnslookup_path
        self.formatter = ColorFormatter()

    def extract_ips(self, output: str) -> List[str]:
        """
        Extracts IP addresses from dnslookup command output.
        
        Args:
            output: Raw output string from dnslookup command
            
        Returns:
            List of extracted IP addresses
        """
        ips = []
        for line in output.split("\n"):
            if "IN" in line and "A" in line:
                parts = line.strip().split()
                if len(parts) >= 5 and "A" in parts:
                    ip = parts[-1]
                    if (
                        all(part.isdigit() for part in ip.split("."))
                        and len(ip.split(".")) == 4
                    ):
                        ips.append(ip)
        return ips

    def resolve_domain(self, domain: str, server: str) -> Dict:
        """
        Resolves a domain using specified DNS server multiple times and collects statistics.
        
        Args:
            domain: Domain name to resolve
            server: DNS server to use
            
        Returns:
            Dictionary containing resolution statistics including:
            - Average response time
            - List of response times
            - Unique IP addresses
            - Error counts and messages
        """
        total_time = 0
        times = []
        all_ips = []
        errors = 0
        error_msgs = set()

        try:
            for i in range(CONFIG["NUM_QUERIES"]):
                start_time = time.time()
                try:
                    process = subprocess.run(
                        [self.dnslookup_path, domain, server],
                        capture_output=True,
                        text=True,
                        check=True,
                    )
                    duration = (time.time() - start_time) * 1000
                    times.append(duration)
                    total_time += duration

                    ips = self.extract_ips(process.stdout)
                    all_ips.extend(ips)

                except subprocess.CalledProcessError as e:
                    errors += 1
                    error_msgs.add(str(e))
                    continue

        except Exception as e:
            errors += 1
            error_msgs.add(str(e))

        successful_queries = len(times)
        avg_time = total_time / successful_queries if successful_queries > 0 else 0
        return {
            "avg_time": avg_time,
            "times": times,
            "unique_ips": list(dict.fromkeys(all_ips)),
            "errors": errors,
            "error_rate": (errors / CONFIG["NUM_QUERIES"]) * 100,
            "error_msgs": list(error_msgs),
            "successful_queries": successful_queries,
            "failed_queries": errors,
        }


class FileHandler:
    """
    Handles file I/O operations for reading domain and DNS server lists.
    Skips empty lines and comments (lines starting with #).
    """
    @staticmethod
    def read_file(filename: str) -> List[str]:
        try:
            with open(filename, "r") as f:
                return [
                    line.strip()
                    for line in f
                    if line.strip() and not line.strip().startswith("#")
                ]
        except FileNotFoundError:
            print(ColorFormatter.error(f"Error: {filename} not found"))
            sys.exit(1)


class ResultPrinter:
    """
    Handles formatted console output of DNS resolution results.
    Uses color coding for better readability.
    """
    def __init__(self):
        self.formatter = ColorFormatter()

    def print_header(self):
        print(f"\n{Fore.CYAN}=== DNS Resolution Results ==={Style.RESET_ALL}")
        print(f"{Fore.CYAN}{'=' * 50}{Style.RESET_ALL}")

    def print_results(self, domain: str, server: str, results: Dict):
        print(f"\n{Fore.GREEN}Domain: {self.formatter.highlight(domain)}")
        print(f"{Fore.GREEN}Server: {Fore.BLUE}{server}{Style.RESET_ALL}")
        print(
            f"{Fore.GREEN}Average Response Time: {Fore.MAGENTA}{results['avg_time']:.2f} ms"
        )
        print(
            f"{Fore.GREEN}Query Times: {Fore.LIGHTBLACK_EX}{' '.join(f'{t:.2f}' for t in results['times'])} ms"
        )
        print(f"{Fore.GREEN}IPs:")

        if results["unique_ips"]:
            for ip in results["unique_ips"]:
                print(f"{Fore.WHITE}  - {ip}{Style.RESET_ALL}")
        else:
            print(self.formatter.error("  No IPs found"))

        print(f"{Fore.LIGHTBLACK_EX}{'-' * 60}{Style.RESET_ALL}")


class ExcelReporter:
    """
    Generates detailed Excel reports of DNS resolution performance.
    Includes server summaries with statistics and color-coded performance indicators.
    """
    def __init__(self):
        self.results_data = []

    def add_result(self, domain: str, server: str, results: Dict):
        has_responses = results["successful_queries"] > 0
        error_rate = results["error_rate"]

        self.results_data.append(
            {
                "Server": server,
                "Domain": domain,
                "Response Time (ms)": round(results["avg_time"], 2)
                if has_responses
                else 0,
                "Error Rate (%)": round(error_rate, 1),
                "Successful Queries": results["successful_queries"],
                "Failed Queries": results["failed_queries"],
                "Working": has_responses and error_rate < 100,
                "Error Messages": "; ".join(results["error_msgs"][:3]),
            }
        )

    def save_report(self):
        """
        Generates an Excel report with the following features:
        - Summary statistics for each DNS server
        - Color-coded performance indicators
        - Auto-sized columns
        - Formatted headers
        - Error rate highlighting
        
        Returns:
            String: Name of the generated Excel file
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"dns_server_performance_{timestamp}.xlsx"

        df = pd.DataFrame(self.results_data)

        agg_dict = {
            "Response Time (ms)": ["mean", "min", "max", "std"],
            "Error Rate (%)": ["mean", "max"],
            "Successful Queries": "sum",
            "Failed Queries": "sum",
            "Working": "any",
            "Error Messages": lambda x: "; ".join(set(msg for msg in x if msg))[:200],
        }

        server_summary = df.groupby("Server").agg(agg_dict).round(2)

        server_summary.columns = [
            "Average (ms)",
            "Min (ms)",
            "Max (ms)",
            "Std Dev (ms)",
            "Avg Error Rate (%)",
            "Max Error Rate (%)",
            "Total Successful Queries",
            "Total Failed Queries",
            "Status",
            "Error Messages",
        ]

        server_summary["Status"] = server_summary["Status"].map(
            lambda x: "Working" if x else "Not Working"
        )

        server_summary["_sort_status"] = (server_summary["Status"] != "Working").astype(
            int
        )
        server_summary["_sort_error"] = server_summary["Avg Error Rate (%)"]
        server_summary = server_summary.sort_values(
            ["_sort_status", "_sort_error", "Average (ms)"]
        ).drop(["_sort_status", "_sort_error"], axis=1)

        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            server_summary.to_excel(writer, sheet_name="Server Performance")

            worksheet = writer.sheets["Server Performance"]

            header_font = Font(bold=True)
            header_fill = PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            good_fill = PatternFill(
                start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"
            )
            warn_fill = PatternFill(
                start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"
            )
            error_fill = PatternFill(
                start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"
            )
            failed_fill = PatternFill(
                start_color="EFEBE9", end_color="EFEBE9", fill_type="solid"
            )

            for col_num, value in enumerate(server_summary.columns.values):
                cell = worksheet.cell(row=1, column=col_num + 2)
                cell.value = value
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
            worksheet.cell(row=1, column=1).value = "Server"
            worksheet.cell(row=1, column=1).font = header_font
            worksheet.cell(row=1, column=1).fill = header_fill
            worksheet.cell(row=1, column=1).border = thin_border

            for row_num, (index, row) in enumerate(server_summary.iterrows(), 2):
                error_rate = row["Avg Error Rate (%)"]
                if row["Status"] != "Working":
                    fill_to_use = failed_fill
                elif error_rate <= 10:
                    fill_to_use = good_fill
                elif error_rate <= 50:
                    fill_to_use = warn_fill
                else:
                    fill_to_use = error_fill

                for col_num in range(len(server_summary.columns)):
                    cell = worksheet.cell(row=row_num, column=col_num + 1)
                    cell.fill = fill_to_use
                    cell.border = thin_border

            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except Exception:
                        pass
                adjusted_width = max_length + 2
                worksheet.column_dimensions[column].width = adjusted_width

        return filename


def is_valid_dns_server(server: str) -> bool:
    """
    Validates if the DNS server starts with tls://, https://, quic://, sdns:// or is a raw IP address.
    
    Args:
        server: DNS server string to validate
    
    Returns:
        bool: True if valid, False otherwise
    """
    prefixes = ("tls://", "https://", "quic://", "sdns://")
    ip_pattern = re.compile(r"^(?:[0-9]{1,3}\.){3}[0-9]{1,3}$")
    
    return server.startswith(prefixes) or bool(ip_pattern.match(server))

def main():
    """
    Main execution function that:
    1. Validates required files and dependencies
    2. Reads configuration files
    3. Performs DNS resolution tests
    4. Displays progress and results
    5. Generates Excel report
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    dnslookup_path = os.path.join(script_dir, CONFIG["DNSLOOKUP_EXE"])

    if not os.path.exists(dnslookup_path):
        print(
            ColorFormatter.error(
                f"Error: {CONFIG['DNSLOOKUP_EXE']} not found at: {dnslookup_path}"
            )
        )
        sys.exit(1)

    file_handler = FileHandler()
    domains = file_handler.read_file(CONFIG["DOMAINS_FILE"])
    doh_servers = file_handler.read_file(CONFIG["DNS_SERVERS_FILE"])
    doh_servers = [server for server in doh_servers if is_valid_dns_server(server)]
    
    if not doh_servers:
        print(ColorFormatter.error("Error: No valid DNS servers found."))
        sys.exit(1)

    resolver = DNSResolver(dnslookup_path)
    printer = ResultPrinter()
    excel_reporter = ExcelReporter()

    printer.print_header()

    total_tasks = len(domains) * len(doh_servers)
    progress = ProgressIndicator(total_tasks)
    progress.start()

    with ThreadPoolExecutor(max_workers=CONFIG["MAX_THREADS"]) as executor:
        future_to_domain = {
            executor.submit(resolver.resolve_domain, domain, server): (domain, server)
            for domain in domains
            for server in doh_servers
        }

        for future in as_completed(future_to_domain):
            domain, server = future_to_domain[future]
            try:
                results = future.result()
                printer.print_results(domain, server, results)
                excel_reporter.add_result(domain, server, results)
            except Exception as e:
                print(ColorFormatter.error(f"Error with {domain} using {server}: {e}"))

            progress.update()

    progress.stop()

    excel_file = excel_reporter.save_report()
    print(ColorFormatter.success(f"\nResults saved to: {excel_file}"))


if __name__ == "__main__":
    main()
